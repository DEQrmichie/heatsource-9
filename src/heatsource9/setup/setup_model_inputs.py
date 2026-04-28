from datetime import datetime, timedelta, timezone
from math import ceil
from pathlib import Path
import logging

from openpyxl.utils import datetime as pyxl_datetime

from heatsource9.io.control_file import cf_path, import_control_file
from heatsource9.io.input_files import read_to_dict, write_input
from heatsource9.io.logging_config import configure_logging
from heatsource9.setup.constants import KM_PRECISION, dtype, sheetnames
from heatsource9.setup.headers import (
    headers_accretion,
    headers_bc,
    headers_inflow,
    headers_lccodes,
    headers_lcdata,
    headers_met,
    headers_met_sites,
    headers_morph,
    headers_trib_sites,
)
from heatsource9.setup.input_setup import InputSetup
from heatsource9.setup.site_setup import _validate_site_file_names


logger = logging.getLogger(__name__)


def _parse_date_yyyy_mm_dd(value):
    return datetime.strptime(value.strip(), "%Y-%m-%d").replace(tzinfo=timezone.utc)


def _hourly_range(start_date, end_date, as_excel):
    """"
    Hourly range is inclusive through end date + 24h.
    """

    start_dt = _parse_date_yyyy_mm_dd(start_date)
    end_dt = _parse_date_yyyy_mm_dd(end_date) + timedelta(days=1)

    values = []
    current = start_dt
    while current <= end_dt:
        if as_excel:
            epoch_excel = datetime(year=1899, month=12, day=30, minute=0, second=0, tzinfo=timezone.utc)
            values.append(pyxl_datetime.to_excel(dt=current, epoch=epoch_excel))
        else:
            values.append(current.strftime("%Y-%m-%d %H:%M"))
        current += timedelta(hours=1)
    return values


def _compute_stream_kms(length_km, longsample_m):
    """
    Return a list of stream kilometers sorted from upstream to downstream.
    """
    num_nodes = int(ceil(round(length_km * 1000 / longsample_m, 4))) + 1
    precision_digits = abs(KM_PRECISION.as_tuple().exponent)
    kms = [
        round((float(node) * longsample_m) / 1000, precision_digits) for node in range(0, num_nodes)
    ]
    kms.sort(reverse=True)
    return kms


def _out_name(params, base, use_timestamp):
    base_name = str(params.get(base) or (base + ".csv"))
    if use_timestamp:
        return datetime.now().strftime("%Y-%m-%d_%H%M%S") + "_" + base_name
    return base_name


def _accretion_table(params, kmlist, use_timestamp):
    rows = [[None, None, km, None, None, None] for km in kmlist]
    filename = _out_name(params, "accretionfile", use_timestamp)
    headers = headers_accretion()
    table = (filename, headers, rows, sheetnames["accretionfile"])
    return table


def _bc_table(params, timelist, use_timestamp):
    rows = [[t, None, None] for t in timelist]
    filename = _out_name(params, "bcfile", use_timestamp)
    headers = headers_bc()
    table = (filename, headers, rows, sheetnames["bcfile"])
    return table


def _morph_table(params, kmlist, use_timestamp):
    rows = [[None, None, km] + [None] * 10 for km in kmlist]
    filename = _out_name(params, "morphfile", use_timestamp)
    headers = headers_morph()
    table = (filename, headers, rows, sheetnames["morphfile"])
    return table


def _lcdata_table(params, kmlist, use_timestamp):
    headers = headers_lcdata(params)
    rows = [[None, None, km] + [None] * (len(headers) - 3) for km in kmlist]
    filename = _out_name(params, "lcdatafile", use_timestamp)
    table = (filename, headers, rows, sheetnames["lcdatafile"])
    return table


def _lccode_table(params, use_timestamp):
    headers = headers_lccodes(params)
    if headers == [None]:
        return None
    rows = [[None]]
    filename = _out_name(params, "lccodefile", use_timestamp)
    table = (filename, headers, rows, sheetnames["lccodefile"])
    return table


def _met_tables(params, timelist, use_timestamp):
    tables = []
    met_files = [p.strip() for p in str(params.get("metfiles") or "").split(",") if p.strip()]
    if not met_files:
        return tables
    headers = headers_met(params)
    rows = [[t] + [None] * (len(headers) - 1) for t in timelist]
    for file in met_files:
        if use_timestamp:
            met_filename = datetime.now().strftime("%Y-%m-%d_%H%M%S") + "_" + file
        else:
            met_filename = file
        table = (met_filename, headers, rows, sheetnames["metfiles"])
        tables.append(table)
    return tables


def _get_met_site_params(model_path, params, control_path, ext, csv_mode):
    setup = InputSetup(control_path)
    metfiles = ""
    metkm = ""
    metheights = ""
    written = None
    met_count = int(params.get("metsites") or 0)
    if met_count <= 0:
        result = (metfiles, metkm, metheights, written)
        return result

    headers = headers_met_sites()
    met_path = model_path / params["metsitesfile"]
    if not met_path.exists():
        rows = []
        metfiles_list = []
        for i in range(1, met_count + 1):
            filename = "metsite{0}{1}".format(i, ext)
            metfiles_list.append(filename)
            rows.append([i, None, None, filename, None])
        write_input(
            path=met_path,
            headers=headers,
            rows=rows,
            sheetname=sheetnames["metsitesfile"],
            csv_mode=csv_mode,
        )
        written = met_path
        metfiles = ", ".join(metfiles_list)
        result = (metfiles, metkm, metheights, written)
        return result

    data = read_to_dict(
        path=met_path,
        colnames=headers,
        sheetname=sheetnames["metsitesfile"],
        value_check=setup._validate,
        header_check=setup.validate_headers,
    )
    met_col_ids = list(data.get("COLID", []))
    met_file_names = list(data.get("FILE_NAME", []))
    met_kms = list(data.get("STREAM_KM", []))
    metheight_values = list(data.get("MET_HEIGHT", []))
    if len(met_col_ids) != met_count:
        msg = "{0} must have exactly {1} data rows because metsites = {1} in the control file.".format(
            met_path.name, met_count
        )
        raise ValueError(msg)
    if any(col_id in (None, "") for col_id in met_col_ids):
        msg = "Met files not written because {0} is missing one or more COLID values.".format(
            met_path.name
        )
        logger.warning(msg)
        result = (metfiles, metkm, metheights, written)
        return result
    if any(file_name in (None, "") for file_name in met_file_names):
        msg = "Met files not written because {0} is missing one or more FILE_NAME values.".format(
            met_path.name
        )
        logger.warning(msg)
        result = (metfiles, metkm, metheights, written)
        return result
    _validate_site_file_names(met_path.name, met_file_names)

    rows = list(zip(met_col_ids, met_file_names, met_kms, metheight_values))
    rows.sort(key=lambda row: row[0])
    metfiles_list = []
    km_values = []
    metheights_list = []
    for met_id, name, stream_km, met_height in rows:
        if name not in metfiles_list:
            metfiles_list.append(name)
        km_values.append("" if stream_km is None else str(stream_km))
        metheights_list.append("" if met_height is None else str(met_height))
    metfiles = ", ".join(metfiles_list)
    metkm = ", ".join(km_values)
    metheights = ", ".join(metheights_list)
    result = (metfiles, metkm, metheights, written)
    return result


def _get_trib_site_params(model_path, params, control_path, ext, csv_mode):
    setup = InputSetup(control_path)
    tribfiles_value = ""
    tribkm_value = ""
    written = None
    inflow_count = int(params.get("tribsites") or 0)
    if inflow_count <= 0:
        result = (tribfiles_value, tribkm_value, written)
        return result

    headers = headers_trib_sites()
    trib_path = model_path / params["tribsitesfile"]
    if not trib_path.exists():
        rows = []
        tribfiles = []
        for i in range(1, inflow_count + 1):
            filename = "tribsite{0}{1}".format(i, ext)
            tribfiles.append(filename)
            rows.append([i, None, None, filename])
        write_input(
            path=trib_path,
            headers=headers,
            rows=rows,
            sheetname=sheetnames["tribsitesfile"],
            csv_mode=csv_mode,
        )
        written = trib_path
        tribfiles_value = ", ".join(tribfiles)
        result = (tribfiles_value, tribkm_value, written)
        return result

    data = read_to_dict(
        path=trib_path,
        colnames=headers,
        sheetname=sheetnames["tribsitesfile"],
        value_check=setup._validate,
        header_check=setup.validate_headers,
    )
    trib_col_ids = list(data.get("COLID", []))
    trib_file_names = list(data.get("FILE_NAME", []))
    trib_kms = list(data.get("STREAM_KM", []))
    if len(trib_col_ids) != inflow_count:
        msg = "{0} must have exactly {1} data rows because tribsites = {1} in the control file.".format(
            trib_path.name, inflow_count
        )
        raise ValueError(msg)
    if any(col_id in (None, "") for col_id in trib_col_ids):
        msg = "Tributary files not written because {0} is missing one or more COLID values.".format(
            trib_path.name
        )
        logger.warning(msg)
        result = (tribfiles_value, tribkm_value, written)
        return result
    if any(file_name in (None, "") for file_name in trib_file_names):
        msg = "Tributary files not written because {0} is missing one or more FILE_NAME values.".format(
            trib_path.name
        )
        logger.warning(msg)
        result = (tribfiles_value, tribkm_value, written)
        return result
    _validate_site_file_names(trib_path.name, trib_file_names)

    rows = list(zip(trib_col_ids, trib_file_names, trib_kms))
    rows.sort(key=lambda row: row[0])
    tribfiles = []
    km_values = []
    for trib_id, name, stream_km in rows:
        if name not in tribfiles:
            tribfiles.append(name)
        km_values.append("" if stream_km is None else str(stream_km))
    tribfiles_value = ", ".join(tribfiles)
    tribkm_value = ", ".join(km_values)
    result = (tribfiles_value, tribkm_value, written)
    return result


def _inflow_tables(params, timelist, use_timestamp):
    tables = []
    inflow_sites = params.get("tribsites") or 0
    if inflow_sites <= 0:
        return tables
    inflow_files_value = params.get("tribfiles")
    inflow_files = [p.strip() for p in str(inflow_files_value or "").split(",") if p.strip()]
    headers = headers_inflow(params)
    rows = [[t] + [None] * (len(headers) - 1) for t in timelist]
    for filename in inflow_files:
        if use_timestamp:
            filename = datetime.now().strftime("%Y-%m-%d_%H%M%S") + "_" + filename
        table = (filename, headers, rows, sheetnames["tribfiles"])
        tables.append(table)
    return tables


def setup_mi(model_dir, control_file = None, *, use_timestamp = False, overwrite = False):
    model_path = Path(model_dir).expanduser().resolve()
    configure_logging(model_dir=model_path, overwrite=True)
    try:
        control_path = cf_path(model_path, control_file)

        params = import_control_file(control_path=control_path,
                                     dtype=dtype,
                                     control_sheet=sheetnames["controlfile"])["control_params"]

        inputdir = Path(str(params.get("inputdir") or (model_path / "Inputs"))).expanduser().resolve()
        outputdir = Path(str(params.get("outputdir") or (model_path / "Output"))).expanduser().resolve()
        for p in (inputdir, outputdir):
            try:
                p.mkdir(parents=True, exist_ok=True)
            except PermissionError as exc:
                msg = f"Cannot create folder '{p}'. Check write permissions."
                raise PermissionError(msg) from exc

        length_km = params.get("length")
        longsample = params.get("longsample")
        if length_km is None or longsample is None or length_km <= 0 or longsample <= 0:
            msg = "Control file must define 'length' and 'longsample' before model input setup"
            raise ValueError(msg)

        datastart = params.get("datastart")
        dataend = params.get("dataend")
        if not datastart or not dataend:
            msg = "Control file must define 'datastart' and 'dataend' before model input setup"
            raise ValueError(msg)

        csv_mode = control_path.suffix.lower() == ".csv"
        timelist = _hourly_range(datastart, dataend, as_excel=not csv_mode)
        kmlist = _compute_stream_kms(length_km, longsample)
        ext = ".csv" if csv_mode else ".xlsx"

        written = []
        met_site_result = _get_met_site_params(model_path, params, control_path, ext, csv_mode)
        params["metfiles"], params["metkm"], params["metheights"], met_site_file = met_site_result
        if met_site_file is not None:
            written.append(met_site_file)

        trib_site_result = _get_trib_site_params(model_path, params, control_path, ext, csv_mode)
        params["tribfiles"], params["tribkm"], trib_site_file = trib_site_result
        if trib_site_file is not None:
            written.append(trib_site_file)

        accretion_table = _accretion_table(params, kmlist, use_timestamp)
        bc_table = _bc_table(params, timelist, use_timestamp)
        morph_table = _morph_table(params, kmlist, use_timestamp)
        lcdata_table = _lcdata_table(params, kmlist, use_timestamp)

        tables = [
            accretion_table,
            bc_table,
            morph_table,
            lcdata_table,
        ]

        lccode_table = _lccode_table(params, use_timestamp)
        if lccode_table is not None:
            tables.append(lccode_table)

        met_tables = _met_tables(params, timelist, use_timestamp)
        for table in met_tables:
            tables.append(table)

        inflow_tables = _inflow_tables(params, timelist, use_timestamp)
        for table in inflow_tables:
            tables.append(table)

        for filename, headers, rows, sheet in tables:
            path = inputdir / filename
            if path.exists() and not overwrite:
                written.append(path)
                continue
            write_input(path=path, headers=headers, rows=rows, sheetname=sheet, csv_mode=csv_mode)
            written.append(path)

        return written
    except Exception:
        msg = "Error in setting up model inputs"
        logger.exception(msg)
        raise
