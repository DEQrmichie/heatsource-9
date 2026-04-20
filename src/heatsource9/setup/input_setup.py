from ast import literal_eval
import re

from pathlib import Path
from time import gmtime
from time import strftime
from datetime import timezone

from openpyxl.utils import datetime as pyxl_datetime

from heatsource9.io.input_files import read_to_dict
from heatsource9.setup.constants import drange, dtype, head2var, sheetnames
from heatsource9.setup.headers import (
    headers_accretion,
    headers_bc,
    headers_inflow,
    headers_lccodes,
    headers_lcdata,
    headers_met,
    headers_morph,
)

import logging

logger = logging.getLogger(__name__)


class InputSetup(object):
    """
    The InputSetup class contains methods to read, parameterize, and write Heat Source input files.
    """

    def __init__(self, control_file_path):
        """
        Initialize input file information from the control file path.
        """
        p = Path(control_file_path).expanduser().resolve()
        self.model_dir = str(p.parent)
        self.control_file = str(p.name)

        # This holds the control file parameters
        self.params= {}

    def _check_range(self, range_key, val, column):
        if range_key is None:
            return
        if not isinstance(val, (int, float)) or isinstance(val, bool):
            return
        lo, hi = drange[range_key]
        if not (lo <= val <= hi):
            msg = f"The {column} value of {val} is out of range [{lo}, {hi}]."
            raise ValueError(msg)

    # ------------------------------------------------------------------
    # Datetime
    # ------------------------------------------------------------------

    def datetime_string(self, dt):
        return strftime("%Y-%m-%d %H:%M", gmtime(dt))

    def datetime_xlsx(self, dt):
        return pyxl_datetime.fromtimestamp(dt, tz=timezone.utc)

    # ------------------------------------------------------------------
    # Validation
    # ------------------------------------------------------------------

    def _validate(self, value, column):
        """
        Checks the data type and range.
        Converts values to the expected data type (dtype) and checks numeric values
        against allowed ranges when a range is defined for the input column.
        Some special logic for Canopy Depth field.
        """
        text = str(column)
        if value is None:
            if text == "CANOPY_DEPTH":
                msg = "CANOPY_DEPTH is required in Land Cover Codes input."
                raise ValueError(msg)
            if self.params.get("lcdatainput") == "Values" and text.startswith("CD_"):
                msg = "When lcdatainput is 'Values', all CD_T*_S* values are required."
                raise ValueError(msg)
            return None
        if isinstance(value, str):
            value = value.strip()
            if value == "":
                if text == "CANOPY_DEPTH":
                    msg = "CANOPY_DEPTH is required in Land Cover Codes input."
                    raise ValueError(msg)
                if self.params.get("lcdatainput") == "Values" and text.startswith("CD_"):
                    msg = "When lcdatainput is 'Values', all CD_T*_S* values are required."
                    raise ValueError(msg)
                return None

        # Remove digits from the column header, e.g. TEMPERATURE2 -> TEMPERATURE.
        base = "".join(ch for ch in text if not ch.isdigit())

        key = None
        range_key = base if base in drange else None

        if column in head2var:
            key = head2var[column]
        elif base in head2var:
            key = head2var[base]
        elif base in dtype:
            key = base
        elif "LC" in text:
            # Find the land cover code columns where headers vary by transect and zone.
            key = "lc_code"
            # Land cover code columns are categorical strings, so no numeric range check.
            range_key = None
        else:
            # Land cover value columns (HT/ELE/LAI/k/CAN/OH/CD) share float dtype.
            # Use the specific header prefix when a direct range exists.
            header_prefix = text.split("_", 1)[0]
            if header_prefix in {"HT", "ELE", "LAI", "k", "CAN", "OH", "CD"}:
                key = "topo"
                if header_prefix in drange:
                    range_key = header_prefix

        if key is None:
            return value
        t = dtype.get(key, str)
        if t is float:
            val = float(value)
            self._check_range(range_key, val, str(column))
            return val
        if t is int:
            val = int(float(value))
            self._check_range(range_key, val, str(column))
            return val
        if t is bool:
            if str(value).title() in ["True", "False"]:
                return literal_eval(str(value).title())
            return bool(value)
        if t in ("datetime", "date"):
            return str(value)
        return str(value)

    def validate_headers(self, filename, colnames, header_row):
        """Validate header count and header names (turned off right now)."""
        ncols_expected = len(colnames)
        ncols_found = len(header_row) if header_row is not None else 0
        file_clean = str(filename).strip()

        if ncols_found < ncols_expected:
            msg = (
                "Input file '{0}' has fewer header columns than expected (expected {1}, found {2}).".format(
                    file_clean, ncols_expected, ncols_found
                )
            )
            raise ValueError(msg)

        check_names = False
        if check_names:
            name_clean = lambda s: re.sub(
                r"^(CLOUDINESS|WIND_SPEED|RELATIVE_HUMIDITY|AIR_TEMPERATURE|FLOW|TEMPERATURE)\d+$",
                r"\1",
                "_".join(str(s).split()).upper(),
            )
            colnames_expected = [name_clean(h) for h in colnames]
            colnames_found = [name_clean(h) for h in header_row[:ncols_expected]]

            if colnames_found != colnames_expected:
                mismatch_i = 0
                for i, (expected_name, found_name) in enumerate(zip(colnames_expected, colnames_found)):
                    if expected_name != found_name:
                        mismatch_i = i
                        break

                raise ValueError(
                    "Input file '{0}' header mismatch at column {1}. Expected {2!r}, found {3!r}.".format(
                        file_clean, mismatch_i + 1, colnames[mismatch_i], header_row[mismatch_i]
                    )
                )

        if ncols_found > ncols_expected:
            msg = (
                "Warning: Input file '{0}' has extra columns beyond the expected number "
                "(expected {1}, found {2}). Extra columns are ignored."
            ).format(file_clean, ncols_expected, ncols_found)
            logger.warning(msg)
    
    # ------------------------------------------------------------------
    # import_* methods
    # ------------------------------------------------------------------

    def dict2list(self, data, colnames, skiprows=0, skipcols=0):
        if not colnames:
            return []
        columns = [[k] + list(data.get(k, [])) for k in colnames]
        rows = list(zip(*columns))
        if skiprows:
            rows = rows[skiprows:]
        if skipcols:
            rows = [row[skipcols:] for row in rows]
        return [list(row) for row in rows]

    def _check_file_exists(self, path):
        if not path.exists():
            msg = (
                "Model input file '{0}' not found. Place the file in the input directory: {1}."
            ).format(path.name, path.parent)
            raise FileNotFoundError(msg)

    def import_lccodes(self):
        """Returns the land cover codes data."""
        path = Path(self.params["inputdir"]) / self.params["lccodefile"]
        self._check_file_exists(path)
        data = read_to_dict(
            path=path,
            colnames=headers_lccodes(self.params),
            sheetname=sheetnames["lccodefile"],
            value_check=self._validate,
            header_check=self.validate_headers,
        )
        # Do some canopy depth validaton that can't be done easily in validate
        nrows = len(data.get("CODE", []))
        heights = data.get("HEIGHT", [])
        for i in range(nrows):
            code = data["CODE"][i]
            h = float(heights[i])
            cd = float(data["CANOPY_DEPTH"][i])
            if cd < 0 or cd > h or (cd == 0 and h > 0):
                msg = (
                    "Canopy depth for code {0} in {1} must be > 0 and <= vegetation height when HEIGHT > 0, "
                    "and must be 0 only when HEIGHT = 0. HEIGHT={2}, CANOPY_DEPTH={3}".format(
                        code, self.params["lccodefile"], h, cd
                    )
                )
                raise ValueError(msg)
            data["CANOPY_DEPTH"][i] = cd
        return data

    def import_lcdata(self, return_list=True, skiprows=1, skipcols=2):
        """Returns the land cover data."""
        headers = headers_lcdata(self.params)
        path = Path(self.params["inputdir"]) / self.params["lcdatafile"]
        self._check_file_exists(path)
        data = read_to_dict(
            path=path,
            colnames=headers,
            sheetname=sheetnames["lcdatafile"],
            value_check=self._validate,
            header_check=self.validate_headers,
        )
        # Do some canopy depth validaton that can't be done easily in validate
        if self.params.get("lcdatainput") == "Values":
            cd_cols = [c for c in headers if isinstance(c, str) and c.startswith("CD_")]
            for cd_col in cd_cols:
                ht_col = "HT_" + cd_col[3:]
                for row_i, (h, cd) in enumerate(zip(data[ht_col], data[cd_col]), start=2):
                    h = float(h)
                    cd = float(cd)
                    if cd < 0 or cd > h or (cd == 0 and h > 0):
                        msg = (
                            "Canopy depth in file {0}, row {1} must be > 0 and <= vegetation height when HEIGHT > 0, "
                            "and must be 0 only when HEIGHT = 0. HEIGHT={2}, CANOPY_DEPTH={3}".format(
                                self.params["lcdatafile"], row_i, h, cd
                            )
                        )
                        raise ValueError(msg)
        if return_list:
            return self.dict2list(data, headers, skiprows, skipcols)
        return data

    def import_inflow(self, return_list=True, skiprows=1, skipcols=1):
        if self.params["tribsites"] <= 0:
            return []
        headers = headers_inflow(self.params)
        filenames = [f.strip() for f in self.params["tribfiles"].split(",") if f.strip()]
        data = []
        for i, filename in enumerate(filenames):
            path = Path(self.params["inputdir"]) / filename
            self._check_file_exists(path)
            site_data = read_to_dict(
                path=path,
                colnames=headers,
                sheetname=sheetnames["tribfiles"],
                value_check=self._validate,
            )
            site_rows = self.dict2list(site_data, headers, skiprows, skipcols)
            if i == 0:
                data = site_rows
            else:
                data = [data[j] + row for j, row in enumerate(site_rows)]
        return data if return_list else data

    def import_morph(self, return_list=False):
        morph_headers = headers_morph()
        path = Path(self.params["inputdir"]) / self.params["morphfile"]
        self._check_file_exists(path)
        data = read_to_dict(
            path=path,
            colnames=morph_headers,
            sheetname=sheetnames["morphfile"],
            value_check=self._validate,
        )
        if return_list:
            return self.dict2list(data, morph_headers, skiprows=1, skipcols=0)
        return data

    def import_met(self, return_list=True, skiprows=1, skipcols=1):
        headers = headers_met(self.params)
        filenames = [f.strip() for f in self.params["metfiles"].split(",") if f.strip()]
        data = []
        for i, filename in enumerate(filenames):
            path = Path(self.params["inputdir"]) / filename
            self._check_file_exists(path)
            site_data = read_to_dict(
                path=path,
                colnames=headers,
                sheetname=sheetnames["metfiles"],
                value_check=self._validate,
            )
            site_rows = self.dict2list(site_data, headers, skiprows, skipcols)
            if i == 0:
                data = site_rows
            else:
                data = [data[j] + row for j, row in enumerate(site_rows)]
        return data if return_list else data

    def import_bc(self, return_list=True, skiprows=1, skipcols=1):
        headers = headers_bc()
        path = Path(self.params["inputdir"]) / self.params["bcfile"]
        self._check_file_exists(path)
        data = read_to_dict(
            path=path,
            colnames=headers,
            sheetname=sheetnames["bcfile"],
            value_check=self._validate,
        )
        if return_list:
            return self.dict2list(data, headers, skiprows, skipcols)
        return data

    def import_accretion(self):
        path = Path(self.params["inputdir"]) / self.params["accretionfile"]
        self._check_file_exists(path)
        return read_to_dict(
            path=path,
            colnames=headers_accretion(),
            sheetname=sheetnames["accretionfile"],
            value_check=self._validate,
        )
