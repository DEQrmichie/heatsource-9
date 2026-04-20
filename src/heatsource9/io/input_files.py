
import csv
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import logging

from openpyxl import Workbook, load_workbook

logger = logging.getLogger(__name__)



def _clean_value(value):
    if isinstance(value, str):
        text = value.strip()
        if text in {"", "None"}:
            return None
        return text
    return value


def read_csv_to_dict(path, colnames, value_check = None, header_check = None):
    """
    Reads a comma delimited text file and returns the data
    as a dictionary with the column header as the key.

    The first row in the source file is skipped as the file header.
    Values are normalized so blank strings and "None" are treated as None.
    Some values are validated for the correct data type and value range.
    """
    data = defaultdict(list, {k: [] for k in colnames})
    expected_cols = len(colnames)
    extra_cols_empty = False
    extra_cols_non_empty = False

    with path.open("r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        header_row = next(reader, [])
        if header_check:
            header_check(path.name, colnames, header_row)
        for row_number, row in enumerate(reader, start=2):
            if not row:
                msg = "Warning: Skipping blank row {0} in input file '{1}'.".format(
                    row_number, path.name
                )
                logger.warning(msg)
                continue
            if all(_clean_value(cell) is None for cell in row):
                msg = "Warning: Skipping blank row {0} in input file '{1}'.".format(
                    row_number, path.name
                )
                logger.warning(msg)
                continue
            if len(row) > expected_cols:
                extras = row[expected_cols:]
                first_non_empty_idx = next(
                    (i for i, v in enumerate(extras) if _clean_value(v) is not None),
                    None,
                )
                if first_non_empty_idx is None:
                    if not extra_cols_empty:
                        msg = (
                            "Warning: Input file '{0}' has trailing empty extra columns beyond the expected "
                            "number (expected {1}, found {2}). Extra columns are ignored."
                        ).format(path.name, expected_cols, len(row))
                        logger.warning(msg)
                        extra_cols_empty = True
                else:
                    if not extra_cols_non_empty:
                        first_col = expected_cols + first_non_empty_idx + 1
                        first_value = extras[first_non_empty_idx]
                        msg = (
                            "Warning: Input file '{0}' has non empty extra columns beyond the expected number "
                            "(expected {1}, found {2}). Extra columns are ignored. First occurrence in column "
                            "{3} row {4} with value {5!r}."
                        ).format(path.name, expected_cols, len(row), first_col, row_number, first_value)
                        logger.warning(msg)
                        extra_cols_non_empty = True
            values = list(row[: expected_cols])
            if len(values) < len(colnames):
                values += [None] * (len(colnames) - len(values))
            for i, key in enumerate(colnames):
                cell = _clean_value(values[i])
                data[key].append(value_check(cell, key) if value_check else cell)
    return data


def read_xlsx_to_dict(
    path,
    colnames,
    sheetname = None,
    value_check = None,
    header_check = None,
):
    """
    Reads an excel file and returns the data
    as a dictionary with the column header as the key.

    The first row in the source worksheet is skipped as the file header.
    Datetime cells are formatted as "YYYY-MM-DD HH:MM".
    Values are normalized so blank strings and "None" are treated as None.
    Some values are validated for the correct data type and value range.
    """
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.worksheets[0] if sheetname is None else wb[sheetname]

    rows = list(ws.iter_rows(values_only=True))
    header_row = list(rows[0]) if rows else []
    if header_check:
        header_check(path.name, colnames, header_row)

    data = defaultdict(list, {k: [] for k in colnames})
    for row_number, row in enumerate(rows[1:], start=2):
        if row is None:
            msg = "Warning: Skipping blank row {0} in input file '{1}'.".format(
                row_number, path.name
            )
            logger.warning(msg)
            continue
        if all(_clean_value(cell) is None for cell in row):
            msg = "Warning: Skipping blank row {0} in input file '{1}'.".format(
                row_number, path.name
            )
            logger.warning(msg)
            continue
        for i, key in enumerate(colnames):
            cell = row[i] if i < len(row) else None
            if isinstance(cell, datetime):
                cell = cell.strftime("%Y-%m-%d %H:%M")
            cell = _clean_value(cell)
            data[key].append(value_check(cell, key) if value_check else cell)
    wb.close()
    return data


def read_csv_to_list(path, skiprows = 0, skipcols = 0):
    """Reads a comma delimited text file into a list of lists indexed by row number."""
    with path.open("r", encoding="utf-8-sig") as f:
        reader = csv.reader(f)
        for _ in range(skiprows):
            next(reader)
        return [row[skipcols:] for row in reader if row]


def read_xlsx_to_list(path, skiprows = 0, skipcols = 0, sheetname = None):
    """Reads an excel worksheet into a list of lists indexed by row number."""
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.worksheets[0] if sheetname is None else wb[sheetname]

    data= []
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r < skiprows or row is None:
            continue
        values = list(row)[skipcols:]
        if values:
            data.append(values)
    wb.close()
    return data


def read_to_dict(
    path,
    colnames,
    sheetname = None,
    value_check = None,
    header_check = None,
):
    """
    Reads a csv or xlsx table and returns the data
    as a dictionary with the column header as the key.

    This is a wrapper for read_csv_to_dict or read_xlsx_to_dict and is
    used to reference the appropriate method based on the input file type.
    """
    if colnames == [None]:
        return {}
    if path.suffix.lower() == ".csv":
        return read_csv_to_dict(path=path, colnames=colnames, value_check=value_check, header_check=header_check)
    return read_xlsx_to_dict(path=path, colnames=colnames, sheetname=sheetname, value_check=value_check, header_check=header_check)


def read_to_list(path, skiprows = 0, skipcols = 0, sheetname = None):
    """
    Reads a csv or xlsx table into a list of lists indexed by row number.

    This is a wrapper for read_csv_to_list or read_xlsx_to_list and is
    used to reference the appropriate method based on the input file type.
    """
    if path.suffix.lower() == ".csv":
        return read_csv_to_list(path=path, skiprows=skiprows, skipcols=skipcols)
    return read_xlsx_to_list(path=path, skiprows=skiprows, skipcols=skipcols, sheetname=sheetname)


def write_input(path, headers, rows, sheetname, csv_mode):
    """
    Write the input files to csv or xlsx.

    Writes the header row first, then writes all data rows in order.
    """
    try:
        path.parent.mkdir(parents=True, exist_ok=True)
    except PermissionError as exc:
        msg = f"Cannot create folder '{path.parent}'. Check write permissions."
        raise PermissionError(msg) from exc

    if csv_mode:
        try:
            with path.open("w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                for row in rows:
                    writer.writerow(["" if v is None else v for v in row])
        except PermissionError as exc:
            msg = f"Cannot write file '{path}'. Check write permissions for this folder."
            raise PermissionError(msg) from exc
        return

    wb = Workbook()
    ws = wb.active
    ws.title = sheetname
    ws.append(headers)
    for row in rows:
        ws.append(row)
    try:
        wb.save(path)
    except PermissionError as exc:
        msg = f"Cannot write file '{path}'. Check write permissions for this folder."
        raise PermissionError(msg) from exc
